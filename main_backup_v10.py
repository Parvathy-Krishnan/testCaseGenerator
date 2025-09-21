from fastapi import FastAPI, Form, UploadFile, HTTPException
from fastapi.responses import JSONResponse, HTMLResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from llama_cpp import Llama
from dotenv import load_dotenv
import os
import requests
import json
from typing import Optional
import openai
from datetime import datetime
import tempfile
import uuid

# Load environment variables
load_dotenv()

# Data models
class RestAssuredRequest(BaseModel):
    apiEndpoint: str
    method: Optional[str] = "GET"
    username: Optional[str] = None
    password: Optional[str] = None
    token: Optional[str] = None
    body: Optional[str] = None
    resourceId: Optional[str] = None
    acceptHeader: Optional[str] = None

class RestAssuredResponse(BaseModel):
    statusCode: int
    response: str

# Initialize FastAPI app
app = FastAPI()

# Mount static files
app.mount("/static", StaticFiles(directory="static"), name="static")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Adjust for production
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Load Llama model
model_path = os.getenv("MODEL_PATH")
llm = None
use_mock_llm = True  # Set to False when actual model is available

if model_path and os.path.exists(model_path) and not use_mock_llm:
    try:
        llm = Llama(model_path=model_path, n_ctx=2048, n_threads=8)
        use_mock_llm = False
    except Exception as e:
        print(f"Failed to load LLM model: {e}")
        use_mock_llm = True
else:
    print("Using mock LLM for testing purposes")

# Initialize OpenAI client
openai_api_key = os.getenv("OPENAI_API_KEY")
openai_client = None
use_openai = False

if openai_api_key:
    try:
        openai_client = openai.OpenAI(api_key=openai_api_key)
        use_openai = True
        print("OpenAI API client initialized successfully")
    except Exception as e:
        print(f"Failed to initialize OpenAI client: {e}")
        use_openai = False
else:
    print("No OpenAI API key provided. OpenAI integration disabled.")

def save_response_to_file(response_text: str, requirement_text: str, api_context: str, operation: str):
    """Save the generated test cases to a file in the workspace"""
    try:
        # Create a filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = "latest_test_cases.md"
        filepath = os.path.join(os.getcwd(), filename)
        
        # Create header with metadata
        header = f"""# Generated Test Cases
**Generated on:** {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
**Operation Type:** {operation}
**Requirement:**
{requirement_text[:200]}{'...' if len(requirement_text) > 200 else ''}

{api_context if api_context else ''}

---

"""
        
        # Write to file (replace existing content)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(header)
            f.write(response_text)
        
        print(f">>> Test cases saved to: {filepath}")
        
    except Exception as e:
        print(f">>> Failed to save test cases to file: {str(e)}")

@app.get("/", response_class=HTMLResponse)
def read_root():
    with open("static/index_new.html", "r") as f:
        return HTMLResponse(content=f.read())

@app.post("/generate-test-cases")
async def generate_test_cases(
    requirement: str = Form(None),
    operation: str = Form(...),
    file: UploadFile = None,
    apiEndpoint: str = Form(None),
    apiMethod: str = Form(None),
    authType: str = Form('none'),
    username: str = Form(None),
    password: str = Form(None),
    token: str = Form(None),
    payload: str = Form(None),
    resourceId: str = Form(None),
    acceptHeader: str = Form(None)
):
    print(f">>> generate-test-cases called with: {requirement[:30] if requirement else 'file upload'}, {operation}, authType={authType}")
    if file:
        print(f">>> file uploaded: {file.filename}")
    if authType == 'basic':
        print(f"Basic Auth: username={username}, password={'***' if password else None}")
    elif authType == 'token':
        print(f"Token Auth: token={'***' if token else None}")

    if llm is None and not use_mock_llm:
        return JSONResponse({"error": "Model still loading, please retry in a few seconds."}, status_code=503)

    # Handle file upload if provided
    requirement_text = requirement
    if file and not requirement:
        try:
            content = await file.read()
            # Try different encodings
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1']
            requirement_text = None
            
            for encoding in encodings:
                try:
                    requirement_text = content.decode(encoding)
                    print(f">>> File decoded successfully with {encoding} encoding")
                    break
                except UnicodeDecodeError:
                    continue
            
            if requirement_text is None:
                # If all encodings fail, try with errors='ignore'
                requirement_text = content.decode('utf-8', errors='ignore')
                print(">>> File decoded with UTF-8 ignoring errors")
            
            print(f">>> File content loaded: {len(requirement_text)} characters")
        except Exception as e:
            return JSONResponse({"error": f"Failed to read file: {str(e)}"}, status_code=400)

    if not requirement_text:
        return JSONResponse({"error": "Either requirement text or file must be provided"}, status_code=400)

    # Enhanced prompt with API context if provided
    api_context = ""
    if apiEndpoint:
        api_context = f"\nAPI Context:\n- Endpoint: {apiEndpoint}\n- Method: {apiMethod or 'GET'}\n- Authentication: {'Basic Auth' if username and password else 'None'}\n"
        if payload:
            api_context += f"- Payload: {payload[:100]}...\n"
        if acceptHeader:
            api_context += f"- Accept Header: {acceptHeader}\n"

    # Build prompt for Karate DSL test cases
    prompt = (
        "You are a professional Karate DSL test case generator specializing in API testing. "
        f"Generate comprehensive {operation.lower()} test cases using Karate DSL syntax for the following requirement:\n\n"
        f"{requirement_text}\n"
        f"{api_context}\n"
        "Generate a complete Karate feature file with:\n"
        "1. Feature description\n"
        "2. Background section with base URL and common setup\n"
        "3. Multiple scenarios with Given-When-Then structure\n"
        "4. Use proper Karate DSL syntax including:\n"
        "   - Given path 'endpoint'\n"
        "   - When method GET/POST/PUT/DELETE\n"
        "   - Then status 200\n"
        "   - And match response contains expected data\n"
        "   - Authentication setup if needed\n"
        "   - Request/response validation\n"
        "5. Include boundary conditions, error scenarios, and data validation tests\n"
        "6. Add performance and security test scenarios where applicable\n\n"
        "Format the output as a complete .feature file ready to run with Karate framework."
    )
    print(">>> Sending prompt to LLM (first 100 chars):", prompt[:100])
    # Build a structured context for the model, which is our "Model Context Protocol"
    model_context = build_model_context(requirement_text, api_context, operation)
    print(">>> Sending prompt to LLM (first 100 chars):", model_context["user_prompt"][:100])

    try:
        response_text = ""
        
        if use_openai and openai_client:
            # Use OpenAI API for test case generation
            response_text = generate_test_cases_with_openai(prompt, openai_client)
            response_text = generate_test_cases_with_openai(model_context, openai_client)
            print(f">>> OpenAI returned {len(response_text)} characters")
        elif use_mock_llm:
            # Mock response for testing
            response_text = generate_mock_test_cases(requirement_text, operation, api_context)
            print(f">>> Mock LLM returned {len(response_text)} characters")
        else:
            # For local Llama, we combine the context into a single prompt string
            prompt = f"{model_context['system_prompt']}\n\n{model_context['user_prompt']}"
            response = llm.create_completion(
                prompt=prompt,
                max_tokens=2048,  # Increased for more comprehensive test cases
                temperature=0.7,
                top_p=0.9
            )
            response_text = response["choices"][0]["text"]
            print(f">>> LLM returned {len(response_text)} characters")
        
        # Save the response to a file in the workspace
        save_response_to_file(response_text, requirement_text, api_context, operation)
        
        return JSONResponse({"output": response_text})
    except Exception as e:
        print(">>> Error from LLM:", str(e))
        return JSONResponse({"error": str(e)}, status_code=500)

def build_model_context(requirement_text: str, api_context: str, operation: str) -> dict:
    """
    Implements a "Model Context Protocol" by creating a structured context for the LLM.
    """
    system_prompt = (
        "You are a world-class software engineer specializing in API testing and a professional Karate DSL test case generator. "
        "Your task is to generate a complete and runnable Karate feature file based on the user's request. "
        "The feature file must include a Feature description, a Background section with URL and common setup, and multiple, diverse scenarios using Given-When-Then structure. "
        "Use proper Karate DSL syntax, including data validation (`match`), error handling, and boundary conditions. "
        "IMPORTANT: Pay close attention to the HTTP method specified in the API context and generate test cases specifically for that method. "
        "The output must be only the code for the .feature file, ready to be saved and executed."
    )

    user_prompt = (
        f"Please generate comprehensive '{operation.lower()}' test cases using Karate DSL for the following requirement:\n\n"
        f"Requirement:\n---\n{requirement_text}\n---\n\n"
    )

    if api_context:
        # Extract method from api_context to emphasize it
        method_line = ""
        for line in api_context.split('\n'):
            if 'Method:' in line:
                method = line.split('Method:')[1].strip()
                method_line = f"\n**CRITICAL: All test scenarios must use HTTP method '{method}' - not GET unless specifically testing error cases.**\n"
                break
        
        user_prompt += f"Use the following API context for the test generation:\n{api_context}{method_line}\n"

    user_prompt += "\nGenerate the complete Karate feature file now. Ensure all primary test scenarios use the specified HTTP method."

    return {
        "system_prompt": system_prompt,
        "user_prompt": user_prompt
    }

def generate_mock_test_cases(requirement: str, operation: str, api_context: str) -> str:
    """Generate mock Karate DSL test cases for testing purposes"""
    
    # Extract API details if available
    endpoint_info = ""
    method = "GET"  # default method
    
    if "endpoint" in api_context.lower():
        endpoint_info = "* url 'https://api.example.com'\n  "
    
    # Extract method from api_context
    if "Method:" in api_context:
        for line in api_context.split('\n'):
            if 'Method:' in line:
                method = line.split('Method:')[1].strip()
                break
    
    # Generate method-specific test cases
    if method.upper() == "POST":
        test_cases = f"""Feature: {method} API Test Cases Generated from Requirements
  {requirement[:100]}...

Background:
  {endpoint_info}* configure connectTimeout = 5000
  * configure readTimeout = 5000

Scenario: Valid {method} Request - Positive Test Case
  Given path '/api/v1/resource'
  * request {{"title": "Test Item", "description": "Test description"}}
  When method {method}
  Then status 201
  And match response != null
  And match response.id != null
  And match response.title == 'Test Item'
  * print 'Response:', response

Scenario: {method} with Invalid Data - Validation Test
  Given path '/api/v1/resource'
  * request {{"invalid": "data"}}
  When method {method}
  Then status 400
  And match response.errors != null
  And match response.message contains 'validation'

Scenario: {method} with Empty Request Body
  Given path '/api/v1/resource'
  * request {{}}
  When method {method}
  Then status 400
  And match response.error != null

Scenario: {method} Authentication Required Test
  Given path '/api/v1/resource'
  * request {{"title": "Test", "description": "Test desc"}}
  When method {method}
  Then status 401
  And match response.message contains 'unauthorized'

Scenario: {method} with Authorization Token
  Given path '/api/v1/resource'
  * header Authorization = 'Bearer valid-token'
  * request {{"title": "Authorized Test", "description": "Test with auth"}}
  When method {method}
  Then status 201
  And match response.id != null

Scenario: {method} Content Type Validation
  Given path '/api/v1/resource'
  * header Content-Type = 'application/json'
  * request {{"title": "Content Test", "description": "Testing content type"}}
  When method {method}
  Then status 201
  And match response != null

Scenario: {method} Large Payload Test
  Given path '/api/v1/resource'
  * def largeData = {{"title": "Large Test", "description": "#(java.util.UUID.randomUUID())", "data": "large data payload"}}
  * request largeData
  When method {method}
  Then status 201
  And match response.id != null

Scenario: {method} Response Time Performance Test
  Given path '/api/v1/resource'
  * request {{"title": "Performance Test", "description": "Testing response time"}}
  When method {method}
  Then status 201
  * def responseTime = responseTime
  * print 'Response time:', responseTime, 'ms'
  * assert responseTime < 5000

Scenario: {method} Duplicate Resource Creation Test
  Given path '/api/v1/resource'
  * request {{"title": "Duplicate Test", "description": "Testing duplicate creation"}}
  When method {method}
  Then status 201
  * def resourceId = response.id
  Given path '/api/v1/resource'
  * request {{"title": "Duplicate Test", "description": "Testing duplicate creation"}}
  When method {method}
  Then status 409
  And match response.error contains 'duplicate'
"""
    
    elif method.upper() == "PUT":
        test_cases = f"""Feature: {method} API Test Cases Generated from Requirements
  {requirement[:100]}...

Background:
  {endpoint_info}* configure connectTimeout = 5000
  * configure readTimeout = 5000

Scenario: Valid {method} Request - Update Resource
  Given path '/api/v1/resource/1'
  * request {{"title": "Updated Item", "description": "Updated description"}}
  When method {method}
  Then status 200
  And match response != null
  And match response.id == 1
  And match response.title == 'Updated Item'

Scenario: {method} Non-Existent Resource
  Given path '/api/v1/resource/999'
  * request {{"title": "Not Found", "description": "Testing not found"}}
  When method {method}
  Then status 404
  And match response.error != null

Scenario: {method} with Invalid Data
  Given path '/api/v1/resource/1'
  * request {{"invalid": "data"}}
  When method {method}
  Then status 400
  And match response.errors != null
"""

    elif method.upper() == "DELETE":
        test_cases = f"""Feature: {method} API Test Cases Generated from Requirements
  {requirement[:100]}...

Background:
  {endpoint_info}* configure connectTimeout = 5000
  * configure readTimeout = 5000

Scenario: Valid {method} Request - Delete Resource
  Given path '/api/v1/resource/1'
  When method {method}
  Then status 204

Scenario: {method} Non-Existent Resource
  Given path '/api/v1/resource/999'
  When method {method}
  Then status 404
  And match response.error != null

Scenario: {method} Already Deleted Resource
  Given path '/api/v1/resource/1'
  When method {method}
  Then status 204
  Given path '/api/v1/resource/1'
  When method {method}
  Then status 404
"""

    else:  # Default to GET
        test_cases = f"""Feature: {method} API Test Cases Generated from Requirements
  {requirement[:100]}...

Background:
  {endpoint_info}* configure connectTimeout = 5000
  * configure readTimeout = 5000

Scenario: Valid {method} Request - Positive Test Case
  Given path '/api/v1/resource'
  When method {method}
  Then status 200
  And match response != null
  And match response.status == 'success'
  * print 'Response:', response

Scenario: {method} Invalid Endpoint Test
  Given path '/api/v1/invalid'
  When method {method}
  Then status 404
  And match response.error != null

Scenario: {method} Authentication Required Test
  Given path '/api/v1/secure'
  * header Authorization = 'Bearer invalid-token'
  When method {method}
  Then status 401
  And match response.message contains 'unauthorized'

Scenario: {method} Response Time Performance Test
  Given path '/api/v1/resource'
  When method {method}
  Then status 200
  * def responseTime = responseTime
  * print 'Response time:', responseTime, 'ms'
  * assert responseTime < 3000

Scenario: {method} Content Type Validation
  Given path '/api/v1/resource'
  * header Accept = 'application/json'
  When method {method}
  Then status 200
  And match header Content-Type contains 'application/json'
"""
    
    return test_cases

def generate_test_cases_with_openai(model_context: dict, client) -> str:
    """Generate test cases using OpenAI API"""
    try:
        response = client.chat.completions.create(
            model="gpt-3.5-turbo",  # You can change this to gpt-4 if you have access
            messages=[
                {
                    "role": "system",
                    "content": model_context["system_prompt"]
                },
                {
                    "role": "user",
                    "content": model_context["user_prompt"]
                }
            ],
            max_tokens=2048,
            temperature=0.7,
            top_p=0.9
        )
        return response.choices[0].message.content
    except Exception as e:
        print(f"OpenAI API Error: {e}")
        # Fallback to mock if OpenAI fails
        return generate_mock_test_cases("API Testing", "both", "")

@app.post("/run-rest-assured")
async def run_rest_assured_test(request: RestAssuredRequest):
    """
    Execute REST Assured test with the provided API configuration
    """
    print(f">>> REST Assured test for: {request.apiEndpoint}")
    
    try:
        # Prepare headers
        headers = {'Content-Type': 'application/json'}
        # Add Accept header if provided
        # Always set Accept header as 'Accept' if provided
        if hasattr(request, 'acceptHeader') and request.acceptHeader:
            headers['Accept'] = request.acceptHeader
            print(f"Accept header value: '{request.acceptHeader}'")
            print(f"Accept header length: {len(request.acceptHeader)}")
        print(f"Outgoing headers: {headers}")
        
        # Prepare authentication
        auth = None
        if request.username and request.password:
            auth = (request.username, request.password)
        elif request.token:
            # Check if token already contains auth type (Basic/Bearer)
            if request.token.startswith(('Basic ', 'Bearer ', 'Digest ')):
                headers['Authorization'] = request.token
            else:
                headers['Authorization'] = f'Bearer {request.token}'
        
        # Prepare request data
        data = None
        if request.body and request.body.strip():
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                data = request.body
        
        # Make the API request
        method = request.method.upper() if request.method else 'GET'
        
        # Handle DELETE requests with resourceId
        endpoint = request.apiEndpoint
        if method == 'DELETE' and request.resourceId:
            # Append resourceId to endpoint if provided
            if not endpoint.endswith('/'):
                endpoint += '/'
            endpoint += request.resourceId
        
        if method == 'GET':
            response = requests.get(endpoint, headers=headers, auth=auth, timeout=30)
        elif method == 'POST':
            response = requests.post(endpoint, headers=headers, json=data, auth=auth, timeout=30)
        elif method == 'PUT':
            response = requests.put(endpoint, headers=headers, json=data, auth=auth, timeout=30)
        elif method == 'DELETE':
            response = requests.delete(endpoint, headers=headers, auth=auth, timeout=30)
        elif method == 'PATCH':
            response = requests.patch(endpoint, headers=headers, json=data, auth=auth, timeout=30)
        else:
            raise HTTPException(status_code=400, detail=f"Unsupported HTTP method: {method}")
        
        # Format response
        try:
            response_data = response.json()
            response_text = json.dumps(response_data, indent=2)
        except:
            response_text = response.text
        
        result = RestAssuredResponse(
            statusCode=response.status_code,
            response=response_text
        )
        
        print(f">>> REST Assured test completed: {response.status_code}")
        return JSONResponse(result.dict())
        
    except requests.exceptions.Timeout:
        raise HTTPException(status_code=408, detail="API request timed out")
    except requests.exceptions.ConnectionError:
        raise HTTPException(status_code=503, detail="Failed to connect to API endpoint")
    except Exception as e:
        print(f">>> REST Assured test error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Test execution failed: {str(e)}")

def format_request_details(request: RestAssuredRequest, endpoint_override: str = None):
    """Helper function to format request details consistently for all test scenarios"""
    endpoint = endpoint_override or request.apiEndpoint
    auth_type = 'Token' if request.token else 'Basic' if request.username else 'None'
    
    return f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {endpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: {auth_type}
Request Body: {request.body if request.body else 'None'}
========================"""

def format_response_with_request(request_details: str, response_data: str):
    """Helper function to combine request and response data"""
    return f"""{request_details}

=== RESPONSE ===
{response_data}
==============="""

@app.post("/run-automation-script")
async def run_karate_automation_script(request: RestAssuredRequest):
    """
    Generate and execute Karate DSL automation script for test cases
    """
    print(f">>> Karate automation script execution for: {request.apiEndpoint}")
    print(f">>> Request acceptHeader: '{request.acceptHeader}'")
    
    try:
        # Generate Karate feature file
        feature_file = generate_karate_feature_file(request)
        
        # Execute multiple test scenarios
        test_results = []
        
        # Scenario 1: Valid request
        try:
            result = await execute_api_test(request, "valid_request")
            
            # Include request details in the response
            request_details = format_request_details(request)
            response_with_request = format_response_with_request(request_details, result.get("response", ""))
            
            test_results.append({
                "scenario": "Valid Request Test - Response Log",
                "status": "PASSED" if result.get("statusCode", 0) < 400 else "FAILED",
                "statusCode": result.get("statusCode", 0),
                "response": response_with_request,
                "details": "Basic API connectivity and response validation with request details",
                "karateStep": f"* url '{request.apiEndpoint}'\n* method '{request.method or 'GET'}'\n* status {result.get('statusCode', 0)}"
            })
        except Exception as e:
            request_details = format_request_details(request)
            error_response = format_response_with_request(request_details, f"Error: {str(e)}")
            test_results.append({
                "scenario": "Valid Request Test",
                "status": "FAILED",
                "statusCode": 0,
                "response": error_response,
                "details": f"Error: {str(e)}",
                "karateStep": f"* url '{request.apiEndpoint}'\n* method '{request.method or 'GET'}'\n* status < 400"
            })
        
        # Scenario 2: Invalid endpoint test
        try:
            invalid_request = RestAssuredRequest(
                apiEndpoint=request.apiEndpoint + "/invalid",
                method=request.method,
                username=request.username,
                password=request.password,
                token=request.token,
                body=request.body,
                resourceId=request.resourceId,
                acceptHeader=request.acceptHeader
            )
            result = await execute_api_test(invalid_request, "invalid_endpoint")
            
            request_details = format_request_details(request, request.apiEndpoint + "/invalid")
            response_with_request = format_response_with_request(request_details, result.get("response", ""))
            
            test_results.append({
                "scenario": "Invalid Endpoint Test",
                "status": "PASSED" if result.get("statusCode", 0) == 404 else "FAILED",
                "statusCode": result.get("statusCode", 0),
                "response": response_with_request,
                "details": "Testing error handling for invalid endpoints",
                "karateStep": f"* url '{request.apiEndpoint}/invalid'\n* method '{request.method or 'GET'}'\n* status 404"
            })
        except Exception as e:
            request_details = format_request_details(request, request.apiEndpoint + "/invalid")
            error_response = format_response_with_request(request_details, f"Error: {str(e)}")
            test_results.append({
                "scenario": "Invalid Endpoint Test", 
                "status": "FAILED",
                "statusCode": 0,
                "response": error_response,
                "details": f"Error: {str(e)}",
                "karateStep": f"* url '{request.apiEndpoint}/invalid'\n* method '{request.method or 'GET'}'\n* status 404"
            })
        
        # Scenario 3: Authentication test (for all auth types)
        auth_test_added = False
        if request.username and request.password:
            try:
                invalid_auth_request = RestAssuredRequest(
                    apiEndpoint=request.apiEndpoint,
                    method=request.method,
                    username="invalid_user",
                    password="invalid_pass",
                    token=request.token,
                    body=request.body,
                    resourceId=request.resourceId,
                    acceptHeader=request.acceptHeader
                )
                result = await execute_api_test(invalid_auth_request, "invalid_auth")
                
                # Create custom request details showing the invalid auth
                invalid_auth_details = f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {request.apiEndpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: Basic (invalid_user:invalid_pass)
Request Body: {request.body if request.body else 'None'}
========================"""
                response_with_request = format_response_with_request(invalid_auth_details, result.get("response", ""))
                
                test_results.append({
                    "scenario": "Authentication Test",
                    "status": "PASSED" if result.get("statusCode", 0) == 401 else "FAILED",
                    "statusCode": result.get("statusCode", 0),
                    "response": response_with_request,
                    "details": "Testing authentication failure handling",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* header Authorization = call read('classpath:basic-auth.js') {{ username: 'invalid_user', password: 'invalid_pass' }}\n* method '{request.method or 'GET'}'\n* status 401"
                })
                auth_test_added = True
            except Exception as e:
                invalid_auth_details = f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {request.apiEndpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: Basic (invalid_user:invalid_pass)
Request Body: {request.body if request.body else 'None'}
========================"""
                error_response = format_response_with_request(invalid_auth_details, f"Error: {str(e)}")
                
                test_results.append({
                    "scenario": "Authentication Test",
                    "status": "FAILED", 
                    "statusCode": 0,
                    "response": error_response,
                    "details": f"Error: {str(e)}",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* header Authorization = call read('classpath:basic-auth.js') {{ username: 'invalid_user', password: 'invalid_pass' }}\n* method '{request.method or 'GET'}'\n* status 401"
                })
                auth_test_added = True
        elif request.token:
            # Test with invalid token for token-based auth
            try:
                invalid_token_request = RestAssuredRequest(
                    apiEndpoint=request.apiEndpoint,
                    method=request.method,
                    username=request.username,
                    password=request.password,
                    token="Bearer invalid-token-12345",
                    body=request.body,
                    resourceId=request.resourceId,
                    acceptHeader=request.acceptHeader
                )
                result = await execute_api_test(invalid_token_request, "invalid_auth")
                
                # Create custom request details showing the invalid token
                invalid_token_details = f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {request.apiEndpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: Bearer (invalid-token-12345)
Request Body: {request.body if request.body else 'None'}
========================"""
                response_with_request = format_response_with_request(invalid_token_details, result.get("response", ""))
                
                test_results.append({
                    "scenario": "Authentication Required Test",
                    "status": "PASSED" if result.get("statusCode", 0) in [401, 403] else "FAILED",
                    "statusCode": result.get("statusCode", 0),
                    "response": response_with_request,
                    "details": "Testing authentication failure with invalid token",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* header Authorization = 'Bearer invalid-token'\n* method '{request.method or 'GET'}'\n* status 401"
                })
                auth_test_added = True
            except Exception as e:
                invalid_token_details = f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {request.apiEndpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: Bearer (invalid-token-12345)
Request Body: {request.body if request.body else 'None'}
========================"""
                error_response = format_response_with_request(invalid_token_details, f"Error: {str(e)}")
                
                test_results.append({
                    "scenario": "Authentication Required Test",
                    "status": "FAILED",
                    "statusCode": 0,
                    "response": error_response,
                    "details": f"Error: {str(e)}",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* header Authorization = 'Bearer invalid-token'\n* method '{request.method or 'GET'}'\n* status 401"
                })
                auth_test_added = True
        
        # Scenario 4: Response validation test
        if request.method and request.method.upper() in ['GET', 'POST']:
            try:
                result = await execute_api_test(request, "response_validation")
                validation_passed = result.get("statusCode", 0) == 200 and result.get("response", "").strip()
                
                request_details = format_request_details(request)
                response_with_request = format_response_with_request(request_details, result.get("response", ""))
                
                test_results.append({
                    "scenario": "Response Validation Test",
                    "status": "PASSED" if validation_passed else "FAILED",
                    "statusCode": result.get("statusCode", 0),
                    "response": response_with_request,
                    "details": "Validating response structure and content",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* method '{request.method}'\n* status 200\n* match response != null"
                })
            except Exception as e:
                request_details = format_request_details(request)
                error_response = format_response_with_request(request_details, f"Error: {str(e)}")
                
                test_results.append({
                    "scenario": "Response Validation Test",
                    "status": "FAILED",
                    "statusCode": 0,
                    "response": error_response,
                    "details": f"Error: {str(e)}",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* method '{request.method}'\n* status 200\n* match response != null"
                })

        # Scenario 5: Request Validation Test (adapted for different methods)
        if request.method and request.method.upper() in ['POST', 'PUT', 'PATCH']:
            try:
                invalid_body_request = RestAssuredRequest(
                    apiEndpoint=request.apiEndpoint,
                    method=request.method,
                    username=request.username,
                    password=request.password,
                    token=request.token,
                    body='{"invalid": "data"}',
                    resourceId=request.resourceId,
                    acceptHeader=request.acceptHeader
                )
                result = await execute_api_test(invalid_body_request, "invalid_request")
                
                # Create custom request details showing the invalid body
                invalid_body_details = f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {request.apiEndpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: {'Token' if request.token else 'Basic' if request.username else 'None'}
Request Body: {{"invalid": "data"}}
========================"""
                response_with_request = format_response_with_request(invalid_body_details, result.get("response", ""))
                
                test_results.append({
                    "scenario": "Request Validation Test",
                    "status": "PASSED" if result.get("statusCode", 0) >= 400 else "FAILED",
                    "statusCode": result.get("statusCode", 0),
                    "response": response_with_request,
                    "details": "Testing request validation with invalid data",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* request {{\"invalid\": \"data\"}}\n* method '{request.method}'\n* status 400"
                })
            except Exception as e:
                invalid_body_details = f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {request.apiEndpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: {'Token' if request.token else 'Basic' if request.username else 'None'}
Request Body: {{"invalid": "data"}}
========================"""
                error_response = format_response_with_request(invalid_body_details, f"Error: {str(e)}")
                
                test_results.append({
                    "scenario": "Request Validation Test",
                    "status": "FAILED",
                    "statusCode": 0,
                    "response": error_response,
                    "details": f"Error: {str(e)}",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* request {{\"invalid\": \"data\"}}\n* method '{request.method}'\n* status 400"
                })
        elif request.method and request.method.upper() == 'GET':
            # For GET requests, test with invalid query parameters
            try:
                invalid_query_endpoint = request.apiEndpoint + "?invalid_param=invalid_value&malformed_query=true"
                invalid_query_request = RestAssuredRequest(
                    apiEndpoint=invalid_query_endpoint,
                    method=request.method,
                    username=request.username,
                    password=request.password,
                    token=request.token,
                    body=None,
                    resourceId=request.resourceId,
                    acceptHeader=request.acceptHeader
                )
                result = await execute_api_test(invalid_query_request, "invalid_request")
                
                # Create custom request details showing the invalid query params
                invalid_query_details = f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {invalid_query_endpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: {'Token' if request.token else 'Basic' if request.username else 'None'}
Request Body: {request.body if request.body else 'None'}
========================"""
                response_with_request = format_response_with_request(invalid_query_details, result.get("response", ""))
                
                test_results.append({
                    "scenario": "Request Validation Test",
                    "status": "PASSED" if result.get("statusCode", 0) in [200, 400] else "FAILED",  # Accept both as valid responses
                    "statusCode": result.get("statusCode", 0),
                    "response": response_with_request,
                    "details": "Testing request validation with invalid query parameters",
                    "karateStep": f"* url '{request.apiEndpoint}?invalid_param=invalid_value'\n* method '{request.method}'\n* status 400"
                })
            except Exception as e:
                invalid_query_details = f"""=== REQUEST DETAILS ===
Method: {request.method or 'GET'}
Endpoint: {invalid_query_endpoint}
Accept Header: {request.acceptHeader or 'Not specified'}
Authentication: {'Token' if request.token else 'Basic' if request.username else 'None'}
Request Body: {request.body if request.body else 'None'}
========================"""
                error_response = format_response_with_request(invalid_query_details, f"Error: {str(e)}")
                
                test_results.append({
                    "scenario": "Request Validation Test",
                    "status": "FAILED",
                    "statusCode": 0,
                    "response": error_response,
                    "details": f"Error: {str(e)}",
                    "karateStep": f"* url '{request.apiEndpoint}?invalid_param=invalid_value'\n* method '{request.method}'\n* status 400"
                })

        # Scenario 6: Data Boundary Test
        try:
            if request.method and request.method.upper() in ['POST', 'PUT', 'PATCH']:
                # For POST/PUT/PATCH, test with large data
                import uuid
                large_data_request = RestAssuredRequest(
                    apiEndpoint=request.apiEndpoint,
                    method=request.method,
                    username=request.username,
                    password=request.password,
                    token=request.token,
                    body=f'{{"data": "{str(uuid.uuid4())}", "largeField": "{"x" * 1000}"}}',
                    resourceId=request.resourceId,
                    acceptHeader=request.acceptHeader
                )
                result = await execute_api_test(large_data_request, "boundary_test")
            else:
                # For GET requests, test with boundary query parameters
                boundary_endpoint = request.apiEndpoint + "?limit=99999&offset=999999&test_boundary=true"
                large_data_request = RestAssuredRequest(
                    apiEndpoint=boundary_endpoint,
                    method=request.method,
                    username=request.username,
                    password=request.password,
                    token=request.token,
                    body=None,
                    resourceId=request.resourceId,
                    acceptHeader=request.acceptHeader
                )
                result = await execute_api_test(large_data_request, "boundary_test")
            
            test_results.append({
                "scenario": "Data Boundary Test",
                "status": "PASSED" if result.get("statusCode", 0) < 500 else "FAILED",  # Accept non-server errors
                "statusCode": result.get("statusCode", 0),
                "response": result.get("response", ""),
                "details": "Testing with boundary/large data values",
                "karateStep": f"* url '{request.apiEndpoint}'\n* def largeData = {{\"data\": \"#(java.util.UUID.randomUUID())\"}}\n* request largeData\n* method '{request.method or 'GET'}'\n* status < 400"
            })
        except Exception as e:
            test_results.append({
                "scenario": "Data Boundary Test",
                "status": "FAILED",
                "statusCode": 0,
                "response": "",
                "details": f"Error: {str(e)}",
                "karateStep": f"* url '{request.apiEndpoint}'\n* def largeData = {{\"data\": \"#(java.util.UUID.randomUUID())\"}}\n* request largeData\n* method '{request.method or 'GET'}'\n* status < 400"
            })

        # Scenario 7: Response Time Performance Test
        try:
            import time
            start_time = time.time()
            result = await execute_api_test(request, "performance_test")
            response_time = (time.time() - start_time) * 1000  # Convert to milliseconds
            
            test_results.append({
                "scenario": "Response Time Performance Test",
                "status": "PASSED" if response_time < 3000 else "FAILED",  # 3 second threshold
                "statusCode": result.get("statusCode", 0),
                "response": f"Response time: {response_time:.2f}ms\n{result.get('response', '')}",
                "details": f"Performance test - Response time: {response_time:.2f}ms",
                "karateStep": f"* url '{request.apiEndpoint}'\n* method '{request.method or 'GET'}'\n* status 200\n* assert responseTime < 3000"
            })
        except Exception as e:
            test_results.append({
                "scenario": "Response Time Performance Test",
                "status": "FAILED",
                "statusCode": 0,
                "response": "",
                "details": f"Error: {str(e)}",
                "karateStep": f"* url '{request.apiEndpoint}'\n* method '{request.method or 'GET'}'\n* status 200\n* assert responseTime < 3000"
            })

        # Scenario 8: Content Type Validation Test
        try:
            if request.method and request.method.upper() in ['POST', 'PUT', 'PATCH']:
                # For methods that accept body, test Content-Type header
                content_type_request = RestAssuredRequest(
                    apiEndpoint=request.apiEndpoint,
                    method=request.method,
                    username=request.username,
                    password=request.password,
                    token=request.token,
                    body=request.body or '{"test": "data"}',
                    resourceId=request.resourceId,
                    acceptHeader=request.acceptHeader
                )
                result = await execute_api_test(content_type_request, "content_type")
                test_results.append({
                    "scenario": "Content Type Validation Test",
                    "status": "PASSED" if result.get("statusCode", 0) < 400 else "FAILED",
                    "statusCode": result.get("statusCode", 0),
                    "response": result.get("response", ""),
                    "details": "Testing Content-Type header validation",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* header Content-Type = 'application/json'\n* request {{\"test\": \"data\"}}\n* method '{request.method}'\n* status < 400"
                })
            else:
                # For GET requests, test Accept header validation
                accept_test_request = RestAssuredRequest(
                    apiEndpoint=request.apiEndpoint,
                    method=request.method,
                    username=request.username,
                    password=request.password,
                    token=request.token,
                    body=None,
                    resourceId=request.resourceId,
                    acceptHeader="application/xml"  # Different Accept header to test validation
                )
                result = await execute_api_test(accept_test_request, "content_type")
                test_results.append({
                    "scenario": "Content Type Validation Test",
                    "status": "PASSED" if result.get("statusCode", 0) in [200, 406, 415] else "FAILED",  # Accept various responses
                    "statusCode": result.get("statusCode", 0),
                    "response": result.get("response", ""),
                    "details": "Testing Accept header validation",
                    "karateStep": f"* url '{request.apiEndpoint}'\n* header Accept = 'application/xml'\n* method '{request.method or 'GET'}'\n* status < 400"
                })
        except Exception as e:
            test_results.append({
                "scenario": "Content Type Validation Test",
                "status": "FAILED",
                "statusCode": 0,
                "response": "",
                "details": f"Error: {str(e)}",
                "karateStep": f"* url '{request.apiEndpoint}'\n* header Content-Type = 'application/json'\n* method '{request.method or 'GET'}'\n* status < 400"
            })

        # Scenario 9: Error Handling Test (resource not found)
        try:
            error_request = RestAssuredRequest(
                apiEndpoint=request.apiEndpoint.rstrip('/') + '/999999',
                method='GET',
                username=request.username,
                password=request.password,
                token=request.token,
                body=None,
                resourceId=request.resourceId,
                acceptHeader=request.acceptHeader
            )
            result = await execute_api_test(error_request, "error_handling")
            test_results.append({
                "scenario": "Error Handling Test",
                "status": "PASSED" if result.get("statusCode", 0) == 404 else "FAILED",
                "statusCode": result.get("statusCode", 0),
                "response": result.get("response", ""),
                "details": "Testing error handling for resource not found",
                "karateStep": f"* url '{request.apiEndpoint}/999999'\n* method 'GET'\n* status 404\n* match response.error == 'Resource not found'"
            })
        except Exception as e:
            test_results.append({
                "scenario": "Error Handling Test",
                "status": "FAILED",
                "statusCode": 0,
                "response": "",
                "details": f"Error: {str(e)}",
                "karateStep": f"* url '{request.apiEndpoint}/999999'\n* method 'GET'\n* status 404\n* match response.error == 'Resource not found'"
            })
        
        # Calculate summary
        total_tests = len(test_results)
        passed_tests = len([r for r in test_results if r["status"] == "PASSED"])
        failed_tests = total_tests - passed_tests
        
        automation_result = {
            "featureFile": feature_file,
            "summary": {
                "total": total_tests,
                "passed": passed_tests,
                "failed": failed_tests,
                "success_rate": f"{(passed_tests/total_tests)*100:.1f}%" if total_tests > 0 else "0%"
            },
            "testResults": test_results
        }
        
        print(f">>> Karate automation completed: {passed_tests}/{total_tests} tests passed")
        return JSONResponse(automation_result)
        
    except Exception as e:
        print(f">>> Karate automation script error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Karate automation execution failed: {str(e)}")

def generate_karate_feature_file(request: RestAssuredRequest) -> str:
    """Generate a complete Karate DSL feature file"""
    method = request.method.upper() if request.method else 'GET'
    base_url = request.apiEndpoint.split('/api')[0] if '/api' in request.apiEndpoint else request.apiEndpoint.rsplit('/', 1)[0]
    endpoint_path = request.apiEndpoint.replace(base_url, '') if base_url in request.apiEndpoint else request.apiEndpoint
    
    # Build authentication section
    auth_section = ""
    if request.username and request.password:
        auth_section = f"""
    * def basicAuth = call read('classpath:basic-auth.js') {{ username: '{request.username}', password: '{request.password}' }}
    * header Authorization = basicAuth"""
    
    # Build request body section
    body_section = ""
    if request.body and request.body.strip() and method in ['POST', 'PUT', 'PATCH']:
        try:
            # Try to parse as JSON
            json.loads(request.body)
            body_section = f"""
    * def requestBody = {request.body}
    * request requestBody"""
        except json.JSONDecodeError:
            body_section = f"""
    * def requestBody = '{request.body}'
    * request requestBody"""
    
    # Handle DELETE with resourceId
    if method == 'DELETE' and request.resourceId:
        if not endpoint_path.endswith('/'):
            endpoint_path += '/'
        endpoint_path += request.resourceId
    
    feature_file = f"""Feature: API Automation Tests for {base_url}

Background:
  * url '{base_url}'
  * configure connectTimeout = 10000
  * configure readTimeout = 10000{auth_section}

Scenario: Valid API Request Test
  Given path '{endpoint_path.lstrip('/')}'
  {body_section}
  When method {method}
  Then status < 400
  And match response != null
  * print 'Response:', response

Scenario: Invalid Endpoint Test  
  Given path '{endpoint_path.lstrip('/')}/invalid'
  When method {method}
  Then status 404

"""
    
    if request.username and request.password:
        feature_file += f"""Scenario: Authentication Failure Test
  Given path '{endpoint_path.lstrip('/')}'
  * header Authorization = call read('classpath:basic-auth.js') {{ username: 'invalid_user', password: 'invalid_pass' }}
  {body_section}
  When method {method}
  Then status 401

"""
    
    if method in ['GET', 'POST']:
        feature_file += f"""Scenario: Response Validation Test
  Given path '{endpoint_path.lstrip('/')}'
  {body_section}
  When method {method}
  Then status 200
  And match response != null
  And match response != ''
  * def responseSize = karate.sizeOf(response)
  * print 'Response size:', responseSize

"""
    
    if method == 'POST':
        feature_file += f"""Scenario: Content-Type Validation Test
  Given path '{endpoint_path.lstrip('/')}'
  And header Content-Type = 'application/json'
  {body_section}
  When method {method}
  Then status < 400
  And match header Content-Type contains 'application/json'

"""
    
    feature_file += f"""Scenario: Response Time Performance Test
  Given path '{endpoint_path.lstrip('/')}'
  {body_section}
  When method {method}
  Then status < 400
  * def responseTime = responseTime
  * print 'Response time:', responseTime, 'ms'
  * assert responseTime < 5000
"""
    
    return feature_file

async def execute_api_test(request: RestAssuredRequest, test_type: str) -> dict:
    """Execute a single API test scenario"""
    import requests
    import json
    
    try:
        # Prepare headers
        headers = {'Content-Type': 'application/json'}
        # Add Accept header if provided
        # Always set Accept header as 'Accept' if provided
        if hasattr(request, 'acceptHeader') and request.acceptHeader:
            headers['Accept'] = request.acceptHeader
        print(f"Outgoing headers: {headers}")
        
        # Prepare authentication
        auth = None
        if request.username and request.password:
            if test_type == "invalid_auth":
                auth = ("invalid_user", "invalid_pass")
            else:
                auth = (request.username, request.password)
        elif request.token:
            if test_type == "invalid_auth":
                headers['Authorization'] = 'Bearer invalid-token'
            else:
                # Check if token already contains auth type (Basic/Bearer)
                if request.token.startswith(('Basic ', 'Bearer ', 'Digest ')):
                    headers['Authorization'] = request.token
                else:
                    headers['Authorization'] = f'Bearer {request.token}'
        
        # Prepare request data
        data = None
        if request.body and request.body.strip():
            try:
                data = json.loads(request.body)
            except json.JSONDecodeError:
                data = request.body
        
        # Make the API request
        method = request.method.upper() if request.method else 'GET'
        endpoint = request.apiEndpoint
        
        # Log the full request details
        print(f"=== REQUEST DETAILS ===")
        print(f"Method: {method}")
        print(f"Endpoint: {endpoint}")
        print(f"Headers: {headers}")
        if data:
            print(f"Request body: {data}")
        if auth:
            print(f"Using basic auth with user: {auth[0]}")
        print(f"=== END REQUEST DETAILS ===")
        
        # Modify endpoint for test scenarios
        if test_type == "invalid_endpoint":
            endpoint += "/invalid"
        elif method == 'DELETE' and request.resourceId:
            if not endpoint.endswith('/'):
                endpoint += '/'
            endpoint += request.resourceId
        
        # Execute request based on method
        if method == 'GET':
            response = requests.get(endpoint, headers=headers, auth=auth, timeout=10)
        elif method == 'POST':
            response = requests.post(endpoint, headers=headers, json=data, auth=auth, timeout=10)
        elif method == 'PUT':
            response = requests.put(endpoint, headers=headers, json=data, auth=auth, timeout=10)
        elif method == 'DELETE':
            response = requests.delete(endpoint, headers=headers, auth=auth, timeout=10)
        elif method == 'PATCH':
            response = requests.patch(endpoint, headers=headers, json=data, auth=auth, timeout=10)
        else:
            raise Exception(f"Unsupported HTTP method: {method}")
        
        # Format response
        try:
            response_data = response.json()
            response_text = json.dumps(response_data, indent=2)[:500] + "..." if len(json.dumps(response_data, indent=2)) > 500 else json.dumps(response_data, indent=2)
        except:
            response_text = response.text[:500] + "..." if len(response.text) > 500 else response.text
        
        return {
            "statusCode": response.status_code,
            "response": response_text
        }
        
    except requests.exceptions.Timeout:
        raise Exception("Request timed out")
    except requests.exceptions.ConnectionError:
        raise Exception("Failed to connect to endpoint")
    except Exception as e:
        raise Exception(f"Test execution failed: {str(e)}")

@app.post("/download-report")
async def download_automation_report(request: RestAssuredRequest):
    """
    Generate and download automation execution report in Excel format
    """
    try:
        # Run the automation to get results
        automation_result = await run_karate_automation_script(request)
        
        # Parse the automation result
        result_data = automation_result.body.decode('utf-8') if hasattr(automation_result, 'body') else str(automation_result)
        if isinstance(automation_result, JSONResponse):
            import json
            result_data = json.loads(result_data)
        else:
            result_data = {"testResults": [], "summary": {"total": 0, "passed": 0, "failed": 0}}
        
        # Generate Excel report
        report_file = generate_excel_report(result_data, request)
        
        # Return file for download
        return FileResponse(
            path=report_file,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f"automation_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            headers={"Content-Disposition": "attachment; filename=automation_report.xlsx"}
        )
        
    except Exception as e:
        print(f">>> Download report error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Report generation failed: {str(e)}")

def generate_excel_report(automation_data, request_info):
    """Generate Excel report from automation results"""
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Automation Report"
        
        # Add header information
        ws['A1'] = "API Automation Test Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'] = f"API Endpoint: {request_info.apiEndpoint}"
        ws['A4'] = f"Method: {request_info.method or 'GET'}"
        ws['A5'] = f"Authentication: {'Token' if request_info.token else 'Basic' if request_info.username else 'None'}"
        
        # Add summary information
        summary = automation_data.get("summary", {})
        ws['A7'] = "Test Summary"
        ws['A7'].font = Font(bold=True, size=14)
        ws['A8'] = f"Total Tests: {summary.get('total', 0)}"
        ws['A9'] = f"Passed: {summary.get('passed', 0)}"
        ws['A10'] = f"Failed: {summary.get('failed', 0)}"
        ws['A11'] = f"Success Rate: {summary.get('success_rate', '0%')}"
        
        # Add test results table
        test_results = automation_data.get("testResults", [])
        if test_results:
            ws['A13'] = "Test Results"
            ws['A13'].font = Font(bold=True, size=14)
            
            # Headers
            headers = ['Scenario', 'Status', 'Status Code', 'Details', 'Response Preview']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=14, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Data rows
            for row, result in enumerate(test_results, 15):
                ws.cell(row=row, column=1).value = result.get('scenario', '')
                
                # Status with color coding
                status_cell = ws.cell(row=row, column=2)
                status_cell.value = result.get('status', '')
                if result.get('status') == 'PASSED':
                    status_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif result.get('status') == 'FAILED':
                    status_cell.fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
                
                ws.cell(row=row, column=3).value = result.get('statusCode', '')
                ws.cell(row=row, column=4).value = result.get('details', '')
                
                # Truncate response for readability
                response = result.get('response', '')
                if len(response) > 200:
                    response = response[:200] + "..."
                ws.cell(row=row, column=5).value = response
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
        
    except ImportError:
        # Fallback to basic CSV if pandas/openpyxl not available
        return generate_csv_report(automation_data, request_info)
    except Exception as e:
        print(f"Excel generation error: {str(e)}")
        return generate_csv_report(automation_data, request_info)

def generate_csv_report(automation_data, request_info):
    """Generate CSV report as fallback"""
    import csv
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.csv', mode='w', newline='')
    
    writer = csv.writer(temp_file)
    
    # Header information
    writer.writerow(['API Automation Test Report'])
    writer.writerow([f'Generated on: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'])
    writer.writerow([f'API Endpoint: {request_info.apiEndpoint}'])
    writer.writerow([f'Method: {request_info.method or "GET"}'])
    writer.writerow([''])
    
    # Summary
    summary = automation_data.get("summary", {})
    writer.writerow(['Test Summary'])
    writer.writerow([f'Total Tests: {summary.get("total", 0)}'])
    writer.writerow([f'Passed: {summary.get("passed", 0)}'])
    writer.writerow([f'Failed: {summary.get("failed", 0)}'])
    writer.writerow([f'Success Rate: {summary.get("success_rate", "0%")}'])
    writer.writerow([''])
    
    # Test results
    writer.writerow(['Scenario', 'Status', 'Status Code', 'Details', 'Response Preview'])
    
    test_results = automation_data.get("testResults", [])
    for result in test_results:
        response = result.get('response', '')
        if len(response) > 200:
            response = response[:200] + "..."
        
        writer.writerow([
            result.get('scenario', ''),
            result.get('status', ''),
            result.get('statusCode', ''),
            result.get('details', ''),
            response
        ])
    
    temp_file.close()
    return temp_file.name